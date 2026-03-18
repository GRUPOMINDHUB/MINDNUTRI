from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from pathlib import Path
import tempfile
import os

LOGO_PATH = Path(__file__).parent.parent / "assets" / "logo_mindhub.png"

# Cores Mindhub
COR_HEADER_BG   = "1A1A1A"
COR_HEADER_FG   = "FFFFFF"
COR_SUBHEADER   = "CC0000"
COR_SUBHEADER_FG= "FFFFFF"
COR_ALT1        = "F7F7F7"
COR_ALT2        = "FFFFFF"
COR_TOTAL_BG    = "2D2D2D"
COR_TOTAL_FG    = "FFFFFF"
COR_CUSTO_BG    = "FFF3CD"
COR_BORDA       = "CCCCCC"


def _borda():
    s = Side(style="thin", color=COR_BORDA)
    return Border(left=s, right=s, top=s, bottom=s)


def gerar_ficha_xlsx(dados: dict, caminho_saida: str) -> str:
    """
    Gera a ficha técnica XLSX a partir dos dados coletados pelo agente.

    dados = {
        "nome_prato": str,
        "classificacao": str,
        "codigo": str,
        "estabelecimento": str,
        "peso_porcao_kg": float,
        "ingredientes": [
            {"nome": str, "unidade": str, "peso_bruto": float,
             "peso_liquido": float, "fc": float, "ic": float, "custo_unit": float}
        ],
        "modo_preparo": [str, str, ...],
    }
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Ficha Técnica"
    ws.sheet_view.showGridLines = False

    # Larguras
    for i, w in enumerate([28, 12, 14, 12, 8, 8, 15, 16, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 55

    # Linha 1 — logo + fundo preto
    ws.merge_cells("A1:I1")
    ws["A1"].fill = PatternFill("solid", start_color=COR_HEADER_BG)
    ws["A1"].value = ""
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    if LOGO_PATH.exists():
        try:
            img = XLImage(str(LOGO_PATH))
            img.width, img.height = 140, 40
            img.anchor = "A1"
            ws.add_image(img)
        except Exception:
            ws["A1"].value = "MINDHUB — Ecossistema de Gestão"
            ws["A1"].font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    else:
        ws["A1"].value = "MINDHUB — Ecossistema de Gestão"
        ws["A1"].font = Font(name="Arial", bold=True, size=14, color="FFFFFF")

    # Linha 2 — título
    ws.row_dimensions[2].height = 28
    ws.merge_cells("A2:I2")
    c = ws["A2"]
    c.value = "FICHA TÉCNICA DE PREPARO"
    c.font = Font(name="Arial", bold=True, size=13, color=COR_SUBHEADER_FG)
    c.fill = PatternFill("solid", start_color=COR_SUBHEADER)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Linhas 3–7 — cabeçalho informativo
    for row in range(3, 8):
        ws.row_dimensions[row].height = 22

    peso_porcao = dados.get("peso_porcao_kg", 0.1)
    estabelecimento = dados.get("estabelecimento", "")

    campos = [
        ("A3", "Estabelecimento:", "B3", "C", estabelecimento, None),
        ("D3", "Código:",          "E3", "I", dados.get("codigo", ""), None),
        ("A4", "Nome da preparação:", "B4", "C", dados.get("nome_prato", ""), None),
        ("D4", "Classificação:",   "E4", "I", dados.get("classificacao", ""), None),
        ("A5", "Rendimento total (kg):", "B5", "C", "=SUM(G11:G42)", "#,##0.000"),
        ("D5", "Peso por porção (kg):", "E5", "I", peso_porcao, "#,##0.000"),
        ("A6", "Custo total:",     "B6", "C", "=SUM(I11:I42)", "R$ #,##0.00"),
        ("D6", "Custo por porção:","E6", "I", '=IFERROR(B6/B7,"-")', "R$ #,##0.00"),
        ("A7", "Nº de porções:",   "B7", "C", '=IFERROR(B5/E5,"-")', "#,##0.0"),
        ("D7", "Data:",            "E7", "I", "", None),
    ]

    for lbl_cell, lbl_val, val_cell, val_end_col, val_val, num_fmt in campos:
        row_num = lbl_cell[1:]
        ws[lbl_cell].value = lbl_val
        ws[lbl_cell].font = Font(name="Arial", bold=True, size=9)
        ws[lbl_cell].fill = PatternFill("solid", start_color="EEEEEE")
        ws[lbl_cell].alignment = Alignment(horizontal="left", vertical="center")
        ws[lbl_cell].border = _borda()
        try:
            ws.merge_cells(f"{val_cell}:{val_end_col}{row_num}")
        except Exception:
            pass
        ws[val_cell].value = val_val
        ws[val_cell].font = Font(name="Arial", size=9)
        ws[val_cell].fill = PatternFill("solid", start_color="FFFFFF")
        ws[val_cell].alignment = Alignment(horizontal="left", vertical="center")
        ws[val_cell].border = _borda()
        if num_fmt:
            ws[val_cell].number_format = num_fmt

    # Linha 8 — espaço
    ws.row_dimensions[8].height = 8

    # Linha 9 — subtítulo ingredientes
    ws.row_dimensions[9].height = 20
    ws.merge_cells("A9:I9")
    c = ws["A9"]
    c.value = "INGREDIENTES"
    c.font = Font(name="Arial", bold=True, size=10, color=COR_SUBHEADER_FG)
    c.fill = PatternFill("solid", start_color=COR_SUBHEADER)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = _borda()

    # Linha 10 — cabeçalhos da tabela
    ws.row_dimensions[10].height = 28
    headers = ["INSUMO", "UNIDADE", "PESO BRUTO\n(kg/un)", "PESO LÍQUIDO\n(kg)",
               "FC", "IC", "PESO C/ COCÇÃO\n(kg)", "CUSTO UNIT.\n(R$/kg ou un)", "CUSTO\nTOTAL (R$)"]
    for i, h in enumerate(headers, 1):
        col = get_column_letter(i)
        c = ws[f"{col}10"]
        c.value = h
        c.font = Font(name="Arial", bold=True, size=8, color=COR_HEADER_FG)
        c.fill = PatternFill("solid", start_color=COR_HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _borda()

    # Ingredientes
    ingredientes = dados.get("ingredientes", [])
    for idx, ing in enumerate(ingredientes):
        row = 11 + idx
        ws.row_dimensions[row].height = 18
        bg = COR_ALT1 if idx % 2 == 0 else COR_ALT2

        ws[f"A{row}"] = ing.get("nome", "")
        ws[f"B{row}"] = ing.get("unidade", "kg")
        ws[f"C{row}"] = ing.get("peso_bruto", ing.get("peso_liquido", 0))
        ws[f"D{row}"] = ing.get("peso_liquido", 0)
        ws[f"E{row}"] = ing.get("fc", 1.0)
        ws[f"F{row}"] = ing.get("ic", 1.0)
        ws[f"G{row}"] = f"=D{row}*F{row}"
        ws[f"H{row}"] = ing.get("custo_unit", 0)
        ws[f"I{row}"] = f"=H{row}*D{row}"

        fmt = {"A": (None,"left"),"B":(None,"center"),
               "C":("#,##0.000","center"),"D":("#,##0.000","center"),
               "E":("#,##0.000","center"),"F":("#,##0.000","center"),
               "G":("#,##0.000","center"),"H":("R$ #,##0.00","center"),
               "I":("R$ #,##0.00","center")}
        for col,(nf,al) in fmt.items():
            c = ws[f"{col}{row}"]
            c.font = Font(name="Arial", size=9)
            c.fill = PatternFill("solid", start_color=bg)
            c.alignment = Alignment(horizontal=al, vertical="center")
            c.border = _borda()
            if nf:
                c.number_format = nf

    # Linhas vazias até 42
    last = 10 + len(ingredientes)
    for row in range(last + 1, 43):
        ws.row_dimensions[row].height = 16
        bg = COR_ALT1 if (row - 11) % 2 == 0 else COR_ALT2
        for col in range(1, 10):
            c = ws.cell(row=row, column=col)
            c.value = ""
            c.fill = PatternFill("solid", start_color=bg)
            c.border = _borda()
            c.font = Font(name="Arial", size=9)

    # Linha 43 — total
    ws.row_dimensions[43].height = 22
    ws.merge_cells("A43:H43")
    c = ws["A43"]
    c.value = "CUSTO TOTAL DA PREPARAÇÃO"
    c.font = Font(name="Arial", bold=True, size=10, color=COR_TOTAL_FG)
    c.fill = PatternFill("solid", start_color=COR_TOTAL_BG)
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.border = _borda()
    c2 = ws["I43"]
    c2.value = "=SUM(I11:I42)"
    c2.font = Font(name="Arial", bold=True, size=10, color=COR_TOTAL_FG)
    c2.fill = PatternFill("solid", start_color=COR_TOTAL_BG)
    c2.alignment = Alignment(horizontal="center", vertical="center")
    c2.number_format = "R$ #,##0.00"
    c2.border = _borda()

    # Linha 44 — custo por porção
    ws.row_dimensions[44].height = 22
    ws.merge_cells("A44:H44")
    c = ws["A44"]
    c.value = "CUSTO POR PORÇÃO"
    c.font = Font(name="Arial", bold=True, size=10)
    c.fill = PatternFill("solid", start_color=COR_CUSTO_BG)
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.border = _borda()
    c2 = ws["I44"]
    c2.value = '=IFERROR(B6/B7,"-")'
    c2.font = Font(name="Arial", bold=True, size=10)
    c2.fill = PatternFill("solid", start_color=COR_CUSTO_BG)
    c2.alignment = Alignment(horizontal="center", vertical="center")
    c2.number_format = "R$ #,##0.00"
    c2.border = _borda()

    # Modo de preparo
    ws.row_dimensions[45].height = 8
    ws.row_dimensions[46].height = 20
    ws.merge_cells("A46:I46")
    c = ws["A46"]
    c.value = "MODO DE PREPARO"
    c.font = Font(name="Arial", bold=True, size=10, color=COR_SUBHEADER_FG)
    c.fill = PatternFill("solid", start_color=COR_SUBHEADER)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = _borda()

    for i, passo in enumerate(dados.get("modo_preparo", [])):
        row = 47 + i
        ws.row_dimensions[row].height = 20
        ws.merge_cells(f"A{row}:I{row}")
        c = ws[f"A{row}"]
        c.value = passo
        c.font = Font(name="Arial", size=9)
        c.fill = PatternFill("solid", start_color=COR_ALT1 if i % 2 == 0 else COR_ALT2)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = _borda()

    # Rodapé
    footer_row = 47 + len(dados.get("modo_preparo", [])) + 1
    ws.row_dimensions[footer_row].height = 18
    ws.merge_cells(f"A{footer_row}:I{footer_row}")
    c = ws[f"A{footer_row}"]
    c.value = "Documento gerado pelo Mindnutri — Agente de IA Mindhub"
    c.font = Font(name="Arial", size=8, italic=True, color="888888")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill("solid", start_color=COR_HEADER_BG)
    c.border = _borda()

    wb.save(caminho_saida)
    return caminho_saida
